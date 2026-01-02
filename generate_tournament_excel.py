"""
Generate a Google-Sheets-friendly tournament workbook that matches the API.
Sheets:
- Teams (Team, Pool, Seed)
- Pool_Play (input)
- Championship (input, auto-populates from Standings)
- Bracket (auto from Championship/Standings, in API format)
- Schedule (auto from Pool_Play, in API format)
- Standings (helper)
"""
import random
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


OUTPUT_FILE = Path("tournament-auto.xlsx")
random.seed(42)

POOL_PLAY_SLOTS = [
    ("9:00 AM", ("A1", "A2"), ("C1", "C2")),
    ("9:21 AM", ("B1", "B2"), ("D1", "D2")),
    ("9:42 AM", ("A2", "A3"), ("C1", "C3")),
    ("10:03 AM", ("B2", "B3"), ("D1", "D3")),
    ("10:24 AM", ("A3", "A1"), ("C1", "C4")),
    ("10:45 AM", ("B3", "B1"), ("D1", "D4")),
    ("11:06 AM", ("C2", "C3"), ("D2", "D3")),
    ("11:27 AM", ("C2", "C4"), ("D2", "D4")),
    ("11:48 AM", ("C3", "C4"), ("D3", "D4")),
]

CHAMPIONSHIP_ROWS = [
    {"time": "12:39 PM", "match": "QF1", "team1_pos": "C-1", "team2_pos": "A-3"},
    {"time": "1:00 PM", "match": "QF2", "team1_pos": "D-1", "team2_pos": "B-3"},
    {"time": "1:21 PM", "match": "QF3", "team1_pos": "A-1", "team2_pos": "C-4"},
    {"time": "1:42 PM", "match": "QF4", "team1_pos": "B-1", "team2_pos": "D-4"},
    {"time": "2:03 PM", "match": "SF1", "team1_ref": "G2", "team2_ref": "G3"},
    {"time": "2:24 PM", "match": "EC1", "team1_ref": "H2", "team2_ref": "H3"},
    {"time": "2:45 PM", "match": "SF2", "team1_ref": "G4", "team2_ref": "G5"},
    {"time": "3:06 PM", "match": "EC2", "team1_ref": "H4", "team2_ref": "H5"},
    {"time": "4:09 PM", "match": "EC Final", "team1_ref": "G7", "team2_ref": "G9"},
    {"time": "4:30 PM", "match": "Final", "team1_ref": "G6", "team2_ref": "G8"},
]


def autosize(ws):
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 3, 42))


def team_lookup_formula(seed_code: str) -> str:
    return f'=IFERROR(INDEX(Teams!A:A,MATCH("{seed_code}",Teams!C:C,0)),"{seed_code}")'


def add_overview(wb):
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Metric", "Value"])
    ws.append(["Total Matches", "=COUNTA(Schedule!A2:A200)+COUNTA(Championship!A2:A200)"])
    ws.append(["Pool Play", "=COUNTA(Schedule!A2:A200)"])
    ws.append(["Championship", "=COUNTA(Championship!A2:A200)"])
    ws.append(["Estimated Finish", "4:30 PM"])
    autosize(ws)


def add_teams(wb):
    ws = wb.create_sheet("Teams")
    ws.append(["Team", "Pool", "Seed"])
    pools = {
        "A": ["Wake Forest", "Durham Storm", "Raleigh Raptors"],
        "B": ["Greensboro Grit", "Chapel Hill Blues", "Triad Titans"],
        "C": ["Charlotte Flyers", "Cary Cobras", "Fayetteville Foxes", "Hickory Hounds"],
        "D": ["Asheville Arrows", "Wilmington Waves", "High Point Hawks", "Outer Banks Orcas"],
    }
    for pool, teams in pools.items():
        for idx, team in enumerate(teams, 1):
            ws.append([team, pool, f"{pool}{idx}"])
    autosize(ws)


def add_pool_play(wb):
    ws = wb.create_sheet("Pool_Play")
    ws.append(["Time", "Field", "Team 1", "Score 1", "Team 2", "Score 2", "Winner"])

    def add_match(time_str, field, team1_seed, team2_seed):
        score1 = random.randint(7, 36)
        score2 = random.randint(7, 36)
        ws.append([time_str, field, team_lookup_formula(team1_seed), score1, team_lookup_formula(team2_seed), score2, None])
        r = ws.max_row
        ws.cell(r, 7).value = f'=IF(OR(D{r}="",F{r}=""),"",IF(D{r}=F{r},"Draw",IF(D{r}>F{r},C{r},E{r})))'

    for time_str, m1, m2 in POOL_PLAY_SLOTS:
        add_match(time_str, "Field 1", m1[0], m1[1])
        add_match(time_str, "Field 2", m2[0], m2[1])

    autosize(ws)


def add_standings(wb):
    ws = wb.create_sheet("Standings")
    ws.append(["Pool", "Seed", "Team", "Wins", "Draws", "Points", "SeedNum", "Rank", "PosKey"])

    teams_sheet = wb["Teams"]
    for row in teams_sheet.iter_rows(min_row=2, values_only=True):
        team, pool, seed = row
        ws.append([pool, seed, team, None, None, None, None, None, None])

    last_row = ws.max_row
    pool_range = f"$A$2:$A${last_row}"
    points_range = f"$F$2:$F${last_row}"
    seednum_range = f"$G$2:$G${last_row}"

    for r in range(2, last_row + 1):
        team_cell = f"C{r}"
        pool_cell = f"A{r}"
        # Wins: count occurrences in Pool_Play winners
        ws.cell(r, 4).value = f"=COUNTIF(Pool_Play!$G:$G,{team_cell})"
        # Draws: count draws involving this team
        ws.cell(r, 5).value = (
            f"=COUNTIFS(Pool_Play!$G:$G,\"Draw\",Pool_Play!$C:$C,{team_cell})+"
            f"COUNTIFS(Pool_Play!$G:$G,\"Draw\",Pool_Play!$E:$E,{team_cell})"
        )
        # Points: 3 per win, 1 per draw
        ws.cell(r, 6).value = f"=D{r}*3+E{r}"
        # SeedNum (for tie-breakers)
        ws.cell(r, 7).value = f"=VALUE(MID(B{r},2,2))"
        # Rank within pool
        ws.cell(r, 8).value = (
            f"=1+SUMPRODUCT((({pool_range}={pool_cell})*({points_range}>F{r}))+"
            f"(({pool_range}={pool_cell})*({points_range}=F{r})*({seednum_range}<G{r})))"
        )
        ws.cell(r, 9).value = f"=A{r}&\"-\"&H{r}"

    autosize(ws)


def add_championship(wb):
    ws = wb.create_sheet("Championship")
    ws.append(["Time", "Match", "Team 1", "Score 1", "Team 2", "Score 2", "Winner", "Loser"])

    for entry in CHAMPIONSHIP_ROWS:
        ws.append([entry["time"], entry["match"], None, None, None, None, None, None])
        r = ws.max_row

        if "team1_pos" in entry:
            ws.cell(r, 3).value = f'=IFERROR(INDEX(Standings!C:C,MATCH("{entry["team1_pos"]}",Standings!I:I,0)),"TBD")'
        else:
            ws.cell(r, 3).value = f"=Championship!{entry['team1_ref']}"

        if "team2_pos" in entry:
            ws.cell(r, 5).value = f'=IFERROR(INDEX(Standings!C:C,MATCH("{entry["team2_pos"]}",Standings!I:I,0)),"TBD")'
        else:
            ws.cell(r, 5).value = f"=Championship!{entry['team2_ref']}"

        ws.cell(r, 4).value = random.randint(7, 36)
        ws.cell(r, 6).value = random.randint(7, 36)
        ws.cell(r, 7).value = f'=IF(OR(D{r}="",F{r}=""),"",IF(D{r}=F{r},"Draw",IF(D{r}>F{r},C{r},E{r})))'
        ws.cell(r, 8).value = f'=IF(OR(D{r}="",F{r}=""),"",IF(D{r}=F{r},"Draw",IF(D{r}<F{r},C{r},E{r})))'

    autosize(ws)


def add_bracket(wb):
    ws = wb.create_sheet("Bracket")

    def add_header(title):
        ws.append([title])

    def add_match(team1_formula, team2_formula, score1_formula, score2_formula):
        ws.append([team1_formula, team2_formula, score1_formula, score2_formula])

    add_header("Quarterfinals")
    add_match("=Championship!C2", "=Championship!E2", "=Championship!D2", "=Championship!F2")
    add_match("=Championship!C3", "=Championship!E3", "=Championship!D3", "=Championship!F3")
    add_match("=Championship!C4", "=Championship!E4", "=Championship!D4", "=Championship!F4")
    add_match("=Championship!C5", "=Championship!E5", "=Championship!D5", "=Championship!F5")

    add_header("Semifinals")
    add_match("=Championship!C6", "=Championship!E6", "=Championship!D6", "=Championship!F6")
    add_match("=Championship!C8", "=Championship!E8", "=Championship!D8", "=Championship!F8")

    add_header("Final")
    add_match("=Championship!C11", "=Championship!E11", "=Championship!D11", "=Championship!F11")

    add_header("Consolation Elite")
    add_match("=Championship!C7", "=Championship!E7", "=Championship!D7", "=Championship!F7")
    add_match("=Championship!C9", "=Championship!E9", "=Championship!D9", "=Championship!F9")

    add_header("Elite Consolation Championship")
    add_match("=Championship!C10", "=Championship!E10", "=Championship!D10", "=Championship!F10")

    add_header("Consolation Development")
    add_match(
        '=IFERROR(INDEX(Standings!C:C,MATCH("C-2",Standings!I:I,0)),"TBD")',
        '=IFERROR(INDEX(Standings!C:C,MATCH("D-2",Standings!I:I,0)),"TBD")',
        "",
        "",
    )
    add_match(
        '=IFERROR(INDEX(Standings!C:C,MATCH("C-3",Standings!I:I,0)),"TBD")',
        '=IFERROR(INDEX(Standings!C:C,MATCH("D-3",Standings!I:I,0)),"TBD")',
        "",
        "",
    )
    add_match(
        '=IFERROR(INDEX(Standings!C:C,MATCH("C-4",Standings!I:I,0)),"TBD")',
        '=IFERROR(INDEX(Standings!C:C,MATCH("D-4",Standings!I:I,0)),"TBD")',
        "",
        "",
    )

    autosize(ws)


def add_schedule(wb):
    ws = wb.create_sheet("Schedule")
    ws.append(["Time", "F1 Team1", "F1 Score", "F1 Team2", "F2 Team1", "F2 Score", "F2 Team2"])

    pool_row = 2
    for time_str, _m1, _m2 in POOL_PLAY_SLOTS:
        f1_row = pool_row
        f2_row = pool_row + 1
        ws.append([
            time_str,
            f"=Pool_Play!C{f1_row}",
            f"=IF(OR(Pool_Play!D{f1_row}=\"\",Pool_Play!F{f1_row}=\"\"),\"\",Pool_Play!D{f1_row}&\"-\"&Pool_Play!F{f1_row})",
            f"=Pool_Play!E{f1_row}",
            f"=Pool_Play!C{f2_row}",
            f"=IF(OR(Pool_Play!D{f2_row}=\"\",Pool_Play!F{f2_row}=\"\"),\"\",Pool_Play!D{f2_row}&\"-\"&Pool_Play!F{f2_row})",
            f"=Pool_Play!E{f2_row}",
        ])
        pool_row += 2

    autosize(ws)


def main():
    wb = openpyxl.Workbook()
    add_overview(wb)
    add_teams(wb)
    add_pool_play(wb)
    add_standings(wb)
    add_championship(wb)
    add_bracket(wb)
    add_schedule(wb)
    wb.save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()
